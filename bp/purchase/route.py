import glb.ViewBase as vb
import glb.uc as uc
import bp.purchase.sql as sql
from flask import Blueprint, render_template
from datetime import datetime
from wtforms import DateField, SubmitField
from flask_wtf import FlaskForm
from flask_wtf.file import FileField
from decorators import login_required
bp = Blueprint('purchase', __name__, url_prefix='/purchase')
from openpyxl import load_workbook
import glb.log as log
class UploadForm(FlaskForm):
    file = FileField('')
    submit = SubmitField('上传')
class ItemUpload(uc.Item):
    def __init__(self):
        super().__init__()
        self.form = UploadForm()
        if self.form.validate_on_submit():
            file = self.form.file.data
            self.filename=file.filename
            # file.save('uploads/' + file.filename)
            # file.save(file.filename)
            wb = load_workbook(file)
            sheets = wb.sheetnames
            sheet = wb[sheets[0]]
            # s0 = f"create table tBom_PurchaseOrders (id INT primary key,[编号] varchar(100))"
            # sql.Set_Create(s0)

            # for i in sheet.iter_cols():
            #     print(i)

            calms_index = []
            calms_name = []
            for i in range(1, 1000):
                value = sheet.cell(row=1, column=i).value
                if value is None:
                    break
                value = str(value).replace(' ', '')
                if sql.calms.contain(value) is False:
                    continue
                calms_name.append(value)
                calms_index.append(i)
            for iRow in range(2, 100000):
                if sheet.cell(row=iRow, column=1).value is None:
                    break
                calms_value = []
                for iCalm in calms_index:
                    value = f"'{sheet.cell(row=iRow, column=iCalm).value}'"
                    calms_value.append(value)
                sql.insert(calms_name, calms_value)


    def get_html(self,s_q,i_offset):
        html = f"""	 
          <form method="POST" enctype="multipart/form-data">
              {self.form.csrf_token}
              {self.form.file.label} {self.form.file}
              {self.form.submit}
          </form>"""
        return html

    def get_html_body(self):
        if self.form.validate_on_submit():
            return uc.get_text(f"文件【{self.filename}】上传成功",i=2)

        else:
            return uc.get_text("数据待上传",i=2)


class ItemPN(uc.ItemInputList):
    def __init__(self):
        super().__init__()
        self.list_str= sql.calms.get_list_pn()

class ItemOptionGx(uc.ItemOption):
    def __init__(self):
        super().__init__(["日期查询","编号查询","数据上传"])
    def set_init(self,s_value):
        if s_value=="":
            self.value="日期查询"
        if self.value=="日期查询":
            self.list_child.append(uc.ItemDate())
        elif self.value=="编号查询":
            self.list_child.append(ItemPN())
        else:
            self.upload=ItemUpload()
            self.list_child.append(self.upload)

    def get_html_body(self):
        if self.value=="日期查询":
            selected_date = self.list_child[0].value
            date_select=datetime.strptime(selected_date, "%Y-%m-%d")
            return sql.calms.get_html_date(date_select)
        elif self.value=="编号查询":
            value=self.list_child[0].value
            return sql.calms.get_html_pn(value)
        elif self.value == "数据上传":
            value = self.list_child[0].value
            return self.upload.get_html_body()
        else:
            return self.value


class Guide(uc.Guide):
    def __init__(self):
        super().__init__()
        self.item_gx = ItemOptionGx()
        self.set_add(self.item_gx)
@bp.route('/detail/')
def detail():
    return "detail"

@bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    guide=Guide()
    html=guide.get_html()
    html+=guide.item_gx.get_html_body()
    log.add_log(f"加载界面-采购单,条件：{guide.s_q}")
    return vb.get_view(bp, html)